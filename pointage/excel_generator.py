# =============================================================================
# FICHIER: excel_generator.py - Génération des rapports Excel
# =============================================================================

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from typing import List, Dict

class ExcelGenerator:
    """Générateur de rapports Excel"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
    
    def generate_rapport_reunion(self, reunion_info: Dict, participants: List[Dict], filepath: str) -> bool:
        """Génère un rapport Excel pour une réunion"""
        try:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Rapport de Réunion"
            
            # Configuration des styles
            title_font = Font(name='Arial', size=16, bold=True)
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            cell_font = Font(name='Arial', size=10)
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
            
            # En-tête du rapport
            row = 1
            
            # Titre principal
            self.ws.merge_cells(f'A{row}:G{row}')
            self.ws[f'A{row}'] = "RAPPORT DE PRÉSENCE - RÉUNION"
            self.ws[f'A{row}'].font = title_font
            self.ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 2
            
            # Informations de la réunion
            info_data = [
                ('Titre de la réunion:', reunion_info.get('titre', 'N/A')),
                ('Date et heure:', reunion_info.get('date', 'N/A')),
                ('Lieu:', reunion_info.get('lieu', 'N/A')),
                ('Nombre de participants:', str(len(participants))),
                ('Rapport généré le:', datetime.now().strftime('%d/%m/%Y à %H:%M:%S'))
            ]
            
            for label, value in info_data:
                self.ws[f'A{row}'] = label
                self.ws[f'B{row}'] = value
                self.ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            row += 2
            
            # En-têtes du tableau
            headers = ['Titre', 'Nom', 'Prénom', 'Service', 'Email', 'Téléphone', 'Heure de pointage']
            for col, header in enumerate(headers, 1):
                cell = self.ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Ajustement de la largeur des colonnes
            column_widths = [8, 15, 15, 20, 25, 15, 18]
            for col, width in enumerate(column_widths, 1):
                self.ws.column_dimensions[chr(64 + col)].width = width
            
            row += 1
            
            # Données des participants
            for participant in participants:
                row_data = [
                    participant.get('titre', ''),
                    participant.get('nom', ''),
                    participant.get('prenom', ''),
                    participant.get('service', ''),
                    participant.get('email', ''),
                    participant.get('telephone', ''),
                    participant.get('heure_pointage', '')
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = self.ws.cell(row=row, column=col, value=value)
                    cell.font = cell_font
                    cell.border = border
                    if col == 7:  # Heure de pointage
                        cell.alignment = Alignment(horizontal='center')
                
                row += 1
            
            # Sauvegarde
            self.wb.save(filepath)
            return True
            
        except Exception as e:
            print(f"Erreur lors de la génération du rapport Excel: {e}")
            return False
    
    def get_default_filename(self, reunion_title: str) -> str:
        """Génère un nom de fichier par défaut"""
        safe_title = "".join(c for c in reunion_title if c.isalnum() or c in (' ', '-', '_')).rstrip()
        date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"Rapport_{safe_title}_{date_str}.xlsx"

